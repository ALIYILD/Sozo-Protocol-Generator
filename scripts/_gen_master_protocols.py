"""
Generate per-modality protocol YAMLs + enrich modality knowledge files
from Master with QEEG places.xlsx
"""
import sys, io, os
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import openpyxl

XLSX = r'C:\Users\yildi\OneDrive\Desktop\neuromodulation protocol knowledge base\Master with QEEG places.xlsx'
REF = r'C:\Users\yildi\sozo_generator\data\reference'
MOD_DIR = os.path.join(REF, 'modality_protocols')
MOD_KN = r'C:\Users\yildi\sozo_generator\sozo_knowledge\knowledge\modalities'
CONDITIONS_DIR = r'C:\Users\yildi\sozo_generator\sozo_knowledge\knowledge\conditions'
os.makedirs(MOD_DIR, exist_ok=True)

wb = openpyxl.load_workbook(XLSX, read_only=True)

def get_rows(sheet):
    ws = wb[sheet]
    data = list(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True))
    # find header row (first row with 'Condition')
    header_idx = 0
    for i, r in enumerate(data):
        if r and r[0] == 'Condition':
            header_idx = i
            break
    header = [str(h) if h else f'col{i}' for i, h in enumerate(data[header_idx])]
    return header, [dict(zip(header, r)) for r in data[header_idx+1:] if any(v is not None for v in r)]

def s(v, maxlen=500):
    if v is None: return 'null'
    t = str(v).strip()[:maxlen]
    if not t: return 'null'
    t2 = t.replace('\\','\\\\').replace('"','\\"')
    return '"' + t2 + '"'

def bl(v, indent=4):
    if not v: return 'null'
    t = str(v).strip().replace('"',"'").replace('\n',' ')
    return '>\n' + ' '*indent + t

# Modality sheets and their slugs/names
MODALITY_SHEETS = [
    ('TPS',        'tps',   'Transcranial Pulse Stimulation'),
    ('TMS_rTMS',   'tms',   'Transcranial Magnetic Stimulation / rTMS'),
    ('tDCS',       'tdcs',  'Transcranial Direct Current Stimulation'),
    ('taVNS_tVNS', 'tavns', 'Transcutaneous Auricular Vagus Nerve Stimulation'),
    ('CES',        'ces',   'Cranial Electrotherapy Stimulation'),
    ('tACS',       'tacs',  'Transcranial Alternating Current Stimulation'),
    ('PBM',        'pbm',   'Photobiomodulation / Low-Level Light Therapy'),
    ('PEMF',       'pemf',  'Pulsed Electromagnetic Field Therapy'),
    ('LIFU_tFUS',  'lifu',  'Low-Intensity Focused Ultrasound'),
    ('tRNS',       'trns',  'Transcranial Random Noise Stimulation'),
    ('DBS',        'dbs',   'Deep Brain Stimulation'),
]

PROTOCOL_COLS = [
    'Condition', 'Brain Target / Stimulation Site', 'EEG Position / Coordinates',
    'Protocol Summary', 'Intensity / Energy', 'Frequency', 'Duration per Session',
    'Total Sessions', 'Pulses / Dose per Session', 'Device(s)', 'Electrode / Coil Montage',
    'Regulatory Status', 'Evidence Level', 'Literature Count', 'Key References',
    'Side Effects', 'Notes'
]

# ── Per-modality YAML files ───────────────────────────────────────────────────
all_protocols = {}  # slug -> list of condition protocols

for sheet, slug, name in MODALITY_SHEETS:
    try:
        header, rows = get_rows(sheet)
    except Exception as e:
        print(f'  SKIP {sheet}: {e}')
        continue

    lines = [
        f'# SOZO Brain Center — {name} Protocol Reference',
        f'# Source: Master with QEEG places.xlsx, April 2026',
        f'# {len(rows)} condition protocols',
        '',
        f'modality: {s(slug)}',
        f'name: {s(name)}',
        '',
        'protocols:',
    ]

    protocols = []
    for r in rows:
        cond = r.get('Condition') or r.get(header[0] if header else 'col0')
        if not cond: continue

        cslug = str(cond).lower().strip()
        # make a safe yaml key
        ckey = cslug.replace(' ', '_').replace('/', '_').replace("'","").replace('(','').replace(')','').replace('.','').replace('-','_').replace(',','').replace('__','_')[:40]

        brain_target = r.get('Brain Target / Stimulation Site')
        eeg_pos = r.get('EEG Position / Coordinates')
        protocol_summary = r.get('Protocol Summary')
        intensity = r.get('Intensity / Energy')
        frequency = r.get('Frequency')
        duration = r.get('Duration per Session')
        sessions = r.get('Total Sessions')
        pulses = r.get('Pulses / Dose per Session')
        devices = r.get('Device(s)')
        montage = r.get('Electrode / Coil Montage')
        regulatory = r.get('Regulatory Status')
        evidence = r.get('Evidence Level')
        lit_count = r.get('Literature Count')
        refs = r.get('Key References')
        side_effects = r.get('Side Effects')
        notes = r.get('Notes')

        lines.append(f'  - condition: {s(cond)}')
        lines.append(f'    brain_target: {s(brain_target)}')
        lines.append(f'    eeg_position: {s(eeg_pos)}')
        lines.append(f'    protocol_summary: {bl(protocol_summary)}')
        lines.append(f'    intensity: {s(intensity)}')
        lines.append(f'    frequency: {s(frequency)}')
        lines.append(f'    duration_per_session: {s(duration)}')
        lines.append(f'    total_sessions: {s(sessions)}')
        lines.append(f'    pulses_per_session: {s(pulses)}')
        lines.append(f'    devices: {s(devices)}')
        lines.append(f'    montage: {s(montage)}')
        lines.append(f'    regulatory_status: {s(regulatory)}')
        lines.append(f'    evidence_level: {s(evidence)}')
        lines.append(f'    literature_count: {s(lit_count)}')
        lines.append(f'    key_references: {bl(refs)}')
        lines.append(f'    side_effects: {bl(side_effects)}')
        lines.append(f'    notes: {bl(notes)}')

        protocols.append({'condition': str(cond), 'brain_target': str(brain_target or ''),
                         'eeg_position': str(eeg_pos or ''), 'regulatory': str(regulatory or ''),
                         'evidence': str(evidence or ''), 'lit_count': str(lit_count or '')})

    out_path = os.path.join(MOD_DIR, f'{slug}_protocols.yaml')
    with open(out_path, 'w', encoding='utf-8') as f:
        f.write('\n'.join(lines) + '\n')

    all_protocols[slug] = protocols
    print(f'  {slug}_protocols.yaml: {len(rows)} protocols')

# ── Conditions_QEEG_Map — enrich condition knowledge YAMLs ───────────────────
try:
    _, qeeg_rows = get_rows('Conditions_QEEG_Map')
    SLUG_MAP = {
        'Depression (MDD)': 'depression',
        "Alzheimer's Disease / Dementia": 'alzheimers',
        "Parkinson's Disease": 'parkinsons',
        'Anxiety Disorders (GAD/PD/SA)': 'anxiety',
        'ADHD': 'adhd',
        'Autism Spectrum Disorder (ASD)': 'asd',
        'Chronic Pain': 'chronic_pain',
        'PTSD': 'ptsd',
        'OCD': 'ocd',
        'Multiple Sclerosis (MS)': 'ms',
        'Long COVID / PACS': 'long_covid',
        'Tinnitus': 'tinnitus',
        'Insomnia': 'insomnia',
        'Stroke Rehabilitation': 'stroke_rehab',
        'TBI': 'tbi',
    }

    updated_cond = 0
    for r in qeeg_rows:
        cond = r.get('Condition')
        if not cond: continue
        sl = SLUG_MAP.get(str(cond).strip())
        if not sl: continue

        fp = os.path.join(CONDITIONS_DIR, f'{sl}.yaml')
        if not os.path.exists(fp): continue

        with open(fp, encoding='utf-8') as f:
            content = f.read()
        if 'network_dysfunction_pattern:' in content:
            continue

        symptoms = r.get('Key Symptoms')
        qeeg = r.get('QEEG Patterns')
        electrodes = r.get('Key QEEG Electrode Sites')
        regions = r.get('Affected Brain Regions')
        networks = r.get('Primary Brain Network(s) Disrupted')
        dysfunction = r.get('Network Dysfunction Pattern')
        techniques = r.get('Recommended Neuromodulation Techniques')
        targets = r.get('Primary Stimulation Target(s)')
        rationale = r.get('Stimulation Rationale')

        block = [
            '\n',
            'neuromod_profile:',
            f'  key_symptoms: {bl(symptoms)}',
            f'  qeeg_patterns: {bl(qeeg)}',
            f'  qeeg_electrodes: {s(electrodes)}',
            f'  affected_regions: {bl(regions)}',
            f'  primary_networks_disrupted: {bl(networks)}',
            f'  network_dysfunction_pattern: {bl(dysfunction)}',
            f'  recommended_techniques: {bl(techniques)}',
            f'  primary_stimulation_targets: {s(targets)}',
            f'  stimulation_rationale: {bl(rationale)}',
        ]

        with open(fp, 'a', encoding='utf-8') as f:
            f.write('\n'.join(block) + '\n')
        updated_cond += 1

    print(f'\nCondition knowledge YAMLs enriched with neuromod_profile: {updated_cond}')
except Exception as e:
    print(f'Conditions_QEEG_Map error: {e}')

# ── Conditions_Matrix — comprehensive paper counts ────────────────────────────
try:
    ws_cm = wb['Conditions_Matrix']
    cm_data = list(ws_cm.iter_rows(min_row=1, values_only=True))
    # Find header row
    header_row = None
    for i, r in enumerate(cm_data):
        if r and r[0] == 'Condition':
            header_row = i
            break
    if header_row is not None:
        header = [str(h) if h else f'col{i}' for i, h in enumerate(cm_data[header_row])]
        rows_cm = [dict(zip(header, r)) for r in cm_data[header_row+1:] if r and r[0]]

        lines_cm = [
            '# SOZO Brain Center — Conditions x Modality Master Evidence Matrix',
            '# Source: Master with QEEG places.xlsx, April 2026',
            f'# {len(rows_cm)} conditions x 11 modalities (approximate PubMed/Scholar paper counts)',
            '',
            'conditions:',
        ]

        for r in rows_cm:
            cond = r.get('Condition')
            if not cond or str(cond).startswith('Values'): continue
            ckey = str(cond).lower().replace(' ','_').replace('/','_').replace("'","").replace('(','').replace(')','')[:35]

            def cnt(k):
                v = r.get(k)
                if v is None: return 'null'
                try: return int(float(str(v).replace(',','')))
                except: return 'null'

            best = r.get('Best Modality/Modalities') or r.get('col12') or r.get('col13')
            lines_cm.append(f'  {ckey}:')
            lines_cm.append(f'    condition: {s(cond)}')
            lines_cm.append('    paper_counts:')
            for mod in ['TPS','TMS','tDCS','taVNS','CES','tACS','PBM','PEMF','LIFU','tRNS','DBS']:
                lines_cm.append(f'      {mod.lower().replace("-","_")}: {cnt(mod)}')
            lines_cm.append(f'    best_modality: {s(best)}')

        with open(os.path.join(REF, 'conditions_matrix_master.yaml'), 'w', encoding='utf-8') as f:
            f.write('\n'.join(lines_cm) + '\n')
        print(f'conditions_matrix_master.yaml: {len(rows_cm)} conditions')
except Exception as e:
    print(f'Conditions_Matrix error: {e}')

# ── Brain_Regions sheet — enrich existing brain_regions.yaml ─────────────────
try:
    _, br_rows = get_rows('Brain_Regions')
    # Append to existing brain_regions.yaml as a supplementary section
    br_lines = [
        '\n# ── Supplementary: Brain Regions from Master Protocol Reference ──',
        '# Source: Master with QEEG places.xlsx, April 2026',
        '# Includes: functions, target conditions, modalities, qEEG biomarker relevance',
        '',
        'regions_extended:',
    ]
    for r in br_rows:
        reg = r.get('Brain Region')
        if not reg: continue
        rkey = str(reg).lower().replace(' ','_').replace("'","").replace('/','_').replace('(','').replace(')','')[:30]
        br_lines.append(f'  {rkey}:')
        br_lines.append(f'    region: {s(reg)}')
        br_lines.append(f'    abbreviation: {s(r.get("Abbreviation"))}')
        br_lines.append(f'    eeg_position: {s(r.get("EEG Position"))}')
        br_lines.append(f'    hemisphere: {s(r.get("Hemisphere"))}')
        br_lines.append(f'    depth: {s(r.get("Depth"))}')
        br_lines.append(f'    functions: {bl(r.get("Functions"))}')
        br_lines.append(f'    key_conditions: {s(r.get("Key Conditions"))}')
        br_lines.append(f'    modalities: {s(r.get("Modalities That Can Target It"))}')
        br_lines.append(f'    notes: {bl(r.get("Notes"))}')
        br_lines.append(f'    qeeg_biomarker_relevance: {bl(r.get("QEEG Biomarker Relevance"))}')

    with open(os.path.join(REF, 'brain_regions.yaml'), 'a', encoding='utf-8') as f:
        f.write('\n'.join(br_lines) + '\n')
    print(f'brain_regions.yaml: enriched with {len(br_rows)} extended entries')
except Exception as e:
    print(f'Brain_Regions error: {e}')

print('\nAll done.')
